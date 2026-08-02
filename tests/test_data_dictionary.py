from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
import pandas as pd
import pyreadstat
from typer.testing import CliRunner

from statconvert.backends.arrow_backend import ArrowBackend
from statconvert.cli import app
from statconvert.dataset import Dataset
from statconvert.metadata import DatasetMetadata, VariableMetadata
from statconvert.metadata.dictionary import (
    DICTIONARY_COLUMNS,
    build_data_dictionary,
)


runner = CliRunner()


def _write_labelled_sav(path: Path) -> Path:
    pyreadstat.write_sav(
        pd.DataFrame(
            {
                "status": [1.0, 2.0],
                "score": [10.0, -99.0],
            }
        ),
        path,
        file_label="Survey",
        note=["Imported"],
        column_labels={
            "status": "Participation status",
            "score": "Assessment score",
        },
        variable_value_labels={
            "status": {1.0: "Active", 2.0: "Inactive"},
        },
        missing_ranges={"score": [{"lo": -99.0, "hi": -99.0}]},
    )
    return path


def _rich_dataset(*, label: str = "Embedded survey") -> Dataset:
    metadata = DatasetMetadata(
        source_format="sav",
        source_backend="pyreadstat",
        dataset_label=label,
        notes=["Review note"],
    )
    metadata.add_variable(
        VariableMetadata(
            name="status",
            label=f"{label} status",
            value_labels={2: "Inactive", 1: "Active"},
            missing_values=[-99],
            missing_ranges=[{"low": 90, "high": 99}],
            storage_type="int64",
            display_format="F8.0",
            measure="nominal",
        )
    )
    return Dataset(
        dataframe=pd.DataFrame({"status": [1, 2]}),
        source_format="sav",
        source_file="survey.sav",
        normalized_metadata=metadata,
        metadata_provenance={
            "dataset": "embedded_arrow",
            "columns": {"status": "embedded_arrow"},
        },
    )


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def test_dictionary_builder_is_ordered_readable_and_uses_blanks():
    dataset = _rich_dataset()
    dataset.dataframe["plain"] = ["a", "b"]
    dataset.__post_init__()

    dictionary = build_data_dictionary(dataset)

    assert [row.column_name for row in dictionary.rows] == ["status", "plain"]
    status, plain = dictionary.rows
    assert status.position == 1
    assert status.value_labels == "1 = Active; 2 = Inactive"
    assert status.missing_values == "-99; [90, 99]"
    assert status.value_label_count == 2
    assert status.missing_value_count == 2
    assert status.metadata_source == "embedded_arrow"
    assert plain.variable_label == ""
    assert plain.value_labels == ""
    assert plain.missing_values == ""


def test_metadata_exports_native_dictionary_csv(tmp_path):
    source = _write_labelled_sav(tmp_path / "survey.sav")
    target = tmp_path / "dictionary.csv"

    result = runner.invoke(
        app,
        ["metadata", str(source), "--export-dictionary", str(target)],
    )

    rows = _read_csv(target)
    assert result.exit_code == 0
    assert "Data dictionary written" in result.output
    assert tuple(rows[0]) == DICTIONARY_COLUMNS
    assert [row["column_name"] for row in rows] == ["status", "score"]
    assert rows[0]["position"] == "1"
    assert rows[0]["variable_label"] == "Participation status"
    assert "1.0 = Active" in rows[0]["value_labels"]
    assert rows[0]["dataset_label"] == "Survey"
    assert rows[0]["dataset_notes"] == "Imported"
    assert rows[0]["metadata_source"] == "native_file"


def test_metadata_exports_readable_xlsx_dictionary(tmp_path):
    source = _write_labelled_sav(tmp_path / "survey.sav")
    target = tmp_path / "dictionary.xlsx"

    result = runner.invoke(
        app,
        ["metadata", str(source), "--export-dictionary", str(target)],
    )

    workbook = openpyxl.load_workbook(target, read_only=True, data_only=True)
    assert result.exit_code == 0
    assert workbook.sheetnames == ["Dictionary", "Dataset", "Value Labels"]
    dictionary = workbook["Dictionary"]
    assert dictionary.max_row == 3
    assert [cell.value for cell in dictionary[1]] == list(DICTIONARY_COLUMNS)
    dataset_values = {
        row[0].value: row[1].value
        for row in workbook["Dataset"].iter_rows(min_row=2)
    }
    assert dataset_values["dataset_label"] == "Survey"
    labels = list(workbook["Value Labels"].iter_rows(min_row=2, values_only=True))
    assert ("status", "1.0", "Active") in labels
    workbook.close()


def test_dictionary_export_uses_active_sidecar_metadata(tmp_path):
    source = tmp_path / "plain.csv"
    target = tmp_path / "dictionary.csv"
    pd.DataFrame({"status": [1]}).to_csv(source, index=False)
    dataset = _rich_dataset(label="Sidecar survey")
    dataset.dataframe = pd.DataFrame({"status": [1]})
    dataset.write_sidecar(source)

    result = runner.invoke(
        app,
        ["metadata", str(source), "--export-dictionary", str(target)],
    )

    row = _read_csv(target)[0]
    assert result.exit_code == 0
    assert row["dataset_label"] == "Sidecar survey"
    assert row["variable_label"] == "Sidecar survey status"
    assert row["metadata_source"] == "automatic_sidecar"


def test_dictionary_export_uses_embedded_arrow_and_sidecar_precedence(tmp_path):
    source = tmp_path / "metadata.parquet"
    embedded_target = tmp_path / "embedded.csv"
    sidecar_target = tmp_path / "sidecar.csv"
    ArrowBackend().write(_rich_dataset(), source)
    sidecar = Dataset.sidecar_path(source)
    payload = json.loads(sidecar.read_text(encoding="utf-8"))
    sidecar.unlink()

    embedded = runner.invoke(
        app,
        [
            "metadata",
            str(source),
            "--export-dictionary",
            str(embedded_target),
        ],
    )
    payload["dataset_metadata"]["dataset_label"] = "Sidecar wins"
    payload["columns"][0]["label"] = "Winning label"
    sidecar.write_text(json.dumps(payload), encoding="utf-8")
    sidecar_result = runner.invoke(
        app,
        [
            "metadata",
            str(source),
            "--export-dictionary",
            str(sidecar_target),
        ],
    )

    embedded_row = _read_csv(embedded_target)[0]
    winning_row = _read_csv(sidecar_target)[0]
    assert embedded.exit_code == sidecar_result.exit_code == 0
    assert embedded_row["metadata_source"] == "embedded_arrow"
    assert embedded_row["dataset_label"] == "Embedded survey"
    assert winning_row["metadata_source"] == "automatic_sidecar"
    assert winning_row["dataset_label"] == "Sidecar wins"
    assert winning_row["variable_label"] == "Winning label"


def test_dictionary_output_safety_and_option_validation(tmp_path):
    source = _write_labelled_sav(tmp_path / "survey.sav")
    target = tmp_path / "dictionary.csv"
    command = [
        "metadata",
        str(source),
        "--export-dictionary",
        str(target),
    ]
    assert runner.invoke(app, command).exit_code == 0

    collision = runner.invoke(app, command)
    overwritten = runner.invoke(app, [*command, "--overwrite-dictionary"])
    missing_parent = runner.invoke(
        app,
        [
            "metadata",
            str(source),
            "--export-dictionary",
            str(tmp_path / "missing" / "dictionary.csv"),
        ],
    )
    unsupported = runner.invoke(
        app,
        [
            "metadata",
            str(source),
            "--export-dictionary",
            str(tmp_path / "dictionary.txt"),
        ],
    )
    overwrite_only = runner.invoke(
        app,
        ["metadata", str(source), "--overwrite-dictionary"],
    )

    assert collision.exit_code == 1
    assert "Data dictionary already exists" in collision.output
    assert "Use --overwrite-dictionary" in collision.output
    assert overwritten.exit_code == 0
    assert missing_parent.exit_code == 1
    assert "Parent folder does not exist" in missing_parent.output
    assert "Create the folder first" in missing_parent.output
    assert unsupported.exit_code == 1
    assert "Unsupported dictionary output format: .txt" in unsupported.output
    assert "Use .csv or .xlsx" in unsupported.output
    assert overwrite_only.exit_code == 1
    assert "--overwrite-dictionary requires --export-dictionary" in (
        overwrite_only.output
    )


def test_dictionary_path_cannot_replace_input(tmp_path):
    source = tmp_path / "source.csv"
    pd.DataFrame({"id": [1]}).to_csv(source, index=False)
    before = source.read_bytes()

    result = runner.invoke(
        app,
        ["metadata", str(source), "--export-dictionary", str(source)],
    )

    assert result.exit_code == 1
    assert "would replace the input file" in result.output
    assert source.read_bytes() == before


def test_selected_workbook_dictionary_works_and_ambiguous_input_fails(tmp_path):
    source = tmp_path / "book.xlsx"
    target = tmp_path / "selected.csv"
    with pd.ExcelWriter(source, engine="xlsxwriter") as workbook:
        pd.DataFrame({"id": [1]}).to_excel(
            workbook,
            sheet_name="Data",
            index=False,
        )
        pd.DataFrame({"code": ["A"]}).to_excel(
            workbook,
            sheet_name="Lookup",
            index=False,
        )

    selected = runner.invoke(
        app,
        [
            "metadata",
            str(source),
            "--object",
            "Data",
            "--export-dictionary",
            str(target),
        ],
    )
    ambiguous = runner.invoke(
        app,
        [
            "metadata",
            str(source),
            "--export-dictionary",
            str(tmp_path / "ambiguous.csv"),
        ],
    )

    assert selected.exit_code == 0
    assert [row["column_name"] for row in _read_csv(target)] == ["id"]
    assert ambiguous.exit_code == 1
    assert not (tmp_path / "ambiguous.csv").exists()


def test_dictionary_options_exist_only_on_metadata_command():
    help_result = runner.invoke(app, ["metadata", "--help"])
    assert help_result.exit_code == 0
    assert "--export-dictionary" in help_result.output
    assert "--overwrite-dictionary" in help_result.output

    for command in ("convert", "batch", "transform"):
        result = runner.invoke(app, [command, "--help"])
        assert result.exit_code == 0
        assert "--export-dictionary" not in result.output
        assert "--overwrite-dictionary" not in result.output
